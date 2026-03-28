"""Run pseudoternary_lle + suggest_experiments for all 9 DES systems and write results to Excel."""

from pathlib import Path

import openpyxl
import si_units as si

from pcsaft_quaternary import pseudoternary_lle

DATA = Path(__file__).parent / "data"
OUT = Path(__file__).parent / "out"
OUT.mkdir(exist_ok=True)
CSV_OUT = OUT / "csv"
CSV_OUT.mkdir(exist_ok=True)

pure_jsons = [
    str(DATA / "thiswork2026_pure.json"),
    str(DATA / "water_models.json"),
]

T = 303.15 * si.KELVIN  # 30 °C
P = 1.0 * si.BAR
SOLUTE = "2-phenylethanol"
DILUENT = "water4C_aparicio2007"

# DES systems: (name, HBA, HBD, first_excel_row)
DES_SYSTEMS = [
    ("ThyCar", "thymol", "carvone", 6),
    ("ThyGer", "thymol", "geraniol", 12),
    ("ThyCarvac", "thymol", "carvacrol", 18),
    ("ThyEug", "thymol", "eugenol", 24),
    ("ThyCam", "thymol", "camphor", 30),
    ("CamCar", "camphor", "carvone", 36),
    ("CamGer", "camphor", "geraniol", 42),
    ("CamCarvac", "camphor", "carvacrol", 48),
    ("CamEug", "camphor", "eugenol", 54),
]

results = {}

for name, hba, hbd, start_row in DES_SYSTEMS:
    print(f"\n{'=' * 60}")
    print(f"Running {name}: {hba} + {hbd}")
    print("=" * 60)
    try:
        tie_lines, suggestions = pseudoternary_lle(
            pure_json=pure_jsons,
            T=T,
            P=P,
            solute=SOLUTE,
            solvent1=hba,
            solvent2=hbd,
            diluent=DILUENT,
            solvent_ratio=1.0,
            output=str(OUT / f"LLE_2PE+{name}+water_T30C"),
            suggest_n=5,
            mass_basis=True,
            csv_output=str(CSV_OUT / f"LLE_2PE+{name}+water_T30C"),
        )
        print(f"  Tie-lines found: {len(tie_lines)}")
        print(f"  Suggestions ({len(suggestions)} points):")
        for i, s in enumerate(suggestions):
            z = s.z_pseudo  # (z_aroma, z_DES, z_water) in mass fracs
            print(
                f"    TL{i + 1}: z_aroma={z[0]:.4f}  z_DES={z[1]:.4f}  z_water={z[2]:.4f}"
            )
        results[name] = (suggestions, start_row)
    except Exception as e:
        print(f"  ERROR: {e}")
        results[name] = (None, start_row)

# Write to Excel
print("\n\nWriting results to Excel...")
EXCEL = Path(__file__).parent.parent / "aromas_equilibrios_v3.0.xlsx"
wb = openpyxl.load_workbook(str(EXCEL))
ws = wb["2-phenylethanol_DES"]

for name, (suggestions, start_row) in results.items():
    if suggestions is None:
        print(f"  Skipping {name} (no suggestions)")
        continue
    for i, s in enumerate(suggestions):
        row = start_row + i
        z_aroma, z_des, z_water = s.z_pseudo
        ws.cell(row=row, column=7).value = round(z_aroma, 5)  # G = z_aroma
        ws.cell(row=row, column=8).value = round(z_des, 5)  # H = z_DES
        ws.cell(row=row, column=9).value = round(z_water, 5)  # I = z_water
    print(f"  Written {name}: rows {start_row}–{start_row + len(suggestions) - 1}")

wb.save(str(EXCEL))
print(f"\nSaved: {EXCEL}")
