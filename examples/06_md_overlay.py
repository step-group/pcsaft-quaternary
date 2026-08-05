"""Overlay MARTINI MD tie-lines (Status='OK') on the PC-SAFT pseudoternary diagrams.

The 'Number Density' sheet is in MARTINI beads/nm^3, so bead counts are divided by
BEADS before fractions are formed.  Bead counts come from
MT-4/martini_DES/martini_DES/ff/*.itp; one W bead represents 4 water molecules.
Rows present in the 'Mol Density' sheet are already per-molecule and override it.

Usage:
    uv run examples/06_md_overlay.py                 # all systems, mass fractions
    uv run examples/06_md_overlay.py ThyCam          # one system
    uv run examples/06_md_overlay.py --molar ThyCam  # mole fractions, '_molar' suffix
"""

import csv
import io
from collections import defaultdict
from pathlib import Path

import openpyxl
import si_units as si

from pcsaft_quaternary import pseudoternary_lle

ROOT = Path(__file__).parent.parent
DATA = Path(__file__).parent / "data"
OUT = Path(__file__).parent / "out"
XLSX = ROOT / "tie_lines-MARTINI.xlsx"

# beads per molecule (ff/*.itp); water is the inverse: 1 W bead == 4 molecules
BEADS = {
    "camphor": 3, "carvone": 4, "carvacrol": 4, "eugenol": 5,
    "thymol": 4, "geraniol": 4, "2-phenylethanol": 4, "water": 0.25,
}
MW = {
    "camphor": 152.24, "carvone": 150.22, "carvacrol": 150.22, "eugenol": 164.20,
    "thymol": 150.22, "geraniol": 154.25, "2-phenylethanol": 122.17, "water": 18.02,
}

# short name used in the existing out/ filenames, keyed by (HBA, HBD)
SYSTEM_NAMES = {
    ("thymol", "carvone"): "ThyCar", ("thymol", "geraniol"): "ThyGer",
    ("thymol", "carvacrol"): "ThyCarvac", ("thymol", "eugenol"): "ThyEug",
    ("thymol", "camphor"): "ThyCam", ("camphor", "carvone"): "CamCar",
    ("camphor", "geraniol"): "CamGer", ("camphor", "carvacrol"): "CamCarvac",
    ("camphor", "eugenol"): "CamEug",
}

SOLUTE = "2-phenylethanol"
DILUENT = "water4C_aparicio2007"

# Real experimental tie-lines, mass fractions, verbatim as measured.  One row per
# phase; rows pair up by vial.  Keyed by (HBA, HBD) so more systems can be added.
EXP_CSV = {
    ("thymol", "camphor"): """\
vial,phase,w_2PE,w_thymol,w_camphor,w_water
E0,organic,0.00000,0.49752,0.48608,0.01640
E0,aqueous,0.00000,0.00034,0.00046,0.99920
E1,organic,0.05152,0.47373,0.45813,0.01663
E1,aqueous,0.00114,0.00018,0.00037,0.99831
E2,organic,0.88404,0.01734,0.01708,0.08153
E2,aqueous,0.01608,0.00001,0.00003,0.98388
E3,organic,0.45210,0.25854,0.25371,0.03566
E3,aqueous,0.00974,0.00010,0.00027,0.98989
E4,organic,0.23552,0.37859,0.36244,0.02346
E4,aqueous,0.00537,0.00015,0.00033,0.99416
E5,organic,0.66922,0.13826,0.13154,0.06098
E5,aqueous,0.01262,0.00005,0.00015,0.98718
""",
}


def _pseudo_frac(dens, hba, hbd, mass_basis=True):
    """{component: molecular density} -> (solute, pseudo-solvent, diluent) fractions.

    Fractions are ratios, so any consistent density unit works (mol/L,
    molecules/nm^3, …) as long as every component uses the same one.  With
    ``mass_basis=False`` the molecular amounts are used directly, giving mole
    fractions.
    """
    amt = {k: v * MW[k] for k, v in dens.items()} if mass_basis else dict(dens)
    total = sum(amt.values())
    return (
        amt[SOLUTE] / total,
        (amt[hba] + amt[hbd]) / total,
        amt["water"] / total,
    )


def _from_beads(n_hba, n_hbd, n_water, n_phe, hba, hbd, mass_basis=True):
    """MARTINI beads/nm^3 -> pseudo-ternary fractions."""
    return _pseudo_frac({
        hba: n_hba / BEADS[hba],
        hbd: n_hbd / BEADS[hbd],
        "water": n_water / BEADS["water"],
        SOLUTE: n_phe / BEADS[SOLUTE],
    }, hba, hbd, mass_basis)


def _from_mol_per_l(c_hba, c_hbd, c_water, c_phe, hba, hbd, mass_basis=True):
    """Molecular molar densities (mol/L) -> pseudo-ternary fractions."""
    return _pseudo_frac(
        {hba: c_hba, hbd: c_hbd, "water": c_water, SOLUTE: c_phe},
        hba, hbd, mass_basis,
    )


def read_exp_tie_lines(hba, hbd, mass_basis=True):
    """Experimental tie-lines for one DES, or [] if none were measured.

    Collapses the two DES components onto the pseudo-solvent axis, matching the
    convention used for the model and MD data.  The CSV is in mass fractions;
    with ``mass_basis=False`` each is divided by its molar mass first, giving
    mole fractions.
    """
    text = EXP_CSV.get((hba, hbd))
    if text is None:
        return []

    by_vial = defaultdict(dict)
    for r in csv.DictReader(io.StringIO(text)):
        w = {
            SOLUTE: float(r["w_2PE"]),
            hba: float(r[f"w_{hba}"]),
            hbd: float(r[f"w_{hbd}"]),
            "water": float(r["w_water"]),
        }
        # w_i / M_i is a relative molar amount; _pseudo_frac normalises it
        by_vial[r["vial"]][r["phase"]] = _pseudo_frac(
            {k: v / MW[k] for k, v in w.items()}, hba, hbd, mass_basis
        )
    return [
        {"phase1_pseudo": p["organic"], "phase2_pseudo": p["aqueous"]}
        for _vial, p in sorted(by_vial.items())
    ]


def read_md_tie_lines(path=XLSX, mass_basis=True):
    """{(hba, hbd): [{'phase1_pseudo':…, 'phase2_pseudo':…}, …]} for Status == 'OK'.

    Rows present in the 'Mol Density' sheet are recomputed molecular densities and
    take precedence; everything else falls back to the raw MARTINI bead counts.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    molar = {}
    if "Mol Density" in wb.sheetnames:
        molar = {
            r[0]: r
            for r in wb["Mol Density"].iter_rows(min_row=4, values_only=True)
            if r[0] is not None
        }

    out = defaultdict(list)
    for row in wb["Number Density"].iter_rows(min_row=4, max_row=27, values_only=True):
        if row[6] != "OK":
            continue
        hba, hbd = row[4], row[5]
        src, conv = (molar[row[0]], _from_mol_per_l) if row[0] in molar else (row, _from_beads)
        out[(hba, hbd)].append({
            "phase1_pseudo": conv(*src[7:11], hba, hbd, mass_basis),    # DES-rich
            "phase2_pseudo": conv(*src[12:16], hba, hbd, mass_basis),   # water-rich
        })
    return dict(out)


def main(only=None, mass_basis=True):
    pure_jsons = [str(DATA / "thiswork2026_pure.json"), str(DATA / "water_models.json")]
    md = read_md_tie_lines(mass_basis=mass_basis)
    suffix = "" if mass_basis else "_molar"

    for (hba, hbd), tls in md.items():
        name = SYSTEM_NAMES[(hba, hbd)]
        if only and name not in only:
            continue
        exp = read_exp_tie_lines(hba, hbd, mass_basis=mass_basis)
        basis = "mass" if mass_basis else "mole"
        print(f"\n{name}: {hba} + {hbd} — {len(tls)} MD, {len(exp)} experimental "
              f"tie-line(s), {basis} fractions")
        pseudoternary_lle(
            pure_json=pure_jsons,
            T=303.15 * si.KELVIN,
            P=1.0 * si.BAR,
            solute=SOLUTE,
            solvent1=hba,
            solvent2=hbd,
            diluent=DILUENT,
            solvent_ratio=1.0,
            mass_basis=mass_basis,
            exp_tie_lines=exp or None,          # red circles
            exp_sets=[("MARTINI MD", "#1f77b4", "s", tls)],   # blue squares
            output=str(OUT / f"LLE_2PE+{name}+water_T30C_MD{suffix}"),
        )


def _self_check():
    md = read_md_tie_lines()
    assert sum(len(v) for v in md.values()) == 13, "expected 13 OK tie-lines"
    assert set(md) == {
        ("camphor", "carvone"), ("camphor", "carvacrol"), ("camphor", "eugenol"),
        ("thymol", "camphor"), ("thymol", "carvone"), ("thymol", "geraniol"),
    }
    for tls in md.values():
        for tl in tls:
            for key in ("phase1_pseudo", "phase2_pseudo"):
                assert abs(sum(tl[key]) - 1.0) < 1e-9, f"{key} does not sum to 1"
            # water-rich phase must actually be water-rich
            assert tl["phase2_pseudo"][2] > 0.9
            assert tl["phase1_pseudo"][1] > 0.9   # DES-rich phase is DES-rich

    # thymol:camphor comes from the recomputed 'Mol Density' rows
    thycam = md[("thymol", "camphor")]
    got = sorted(round(tl["phase1_pseudo"][0], 5) for tl in thycam)
    assert got == [0.01054, 0.04163, 0.08233], f"unexpected ThyCam w(2PE): {got}"

    for basis in (True, False):
        exp = read_exp_tie_lines("thymol", "camphor", mass_basis=basis)
        assert len(exp) == 6, f"expected 6 experimental vials, got {len(exp)}"
        for tl in exp:
            for key in ("phase1_pseudo", "phase2_pseudo"):
                assert abs(sum(tl[key]) - 1.0) < 1e-9, f"exp {key} does not sum to 1"
        # organic phase must hold more solute than the aqueous one it coexists with
        assert all(tl["phase1_pseudo"][0] >= tl["phase2_pseudo"][0] for tl in exp)

    # mole basis: same tie-lines, water-heavy on a per-molecule count
    md_x = read_md_tie_lines(mass_basis=False)
    assert sum(len(v) for v in md_x.values()) == 13
    for tls in md_x.values():
        for tl in tls:
            for key in ("phase1_pseudo", "phase2_pseudo"):
                assert abs(sum(tl[key]) - 1.0) < 1e-9, f"mole {key} does not sum to 1"


if __name__ == "__main__":
    import sys

    args = sys.argv[1:]
    molar = "--molar" in args
    _self_check()
    main(only=[a for a in args if not a.startswith("-")] or None, mass_basis=not molar)
